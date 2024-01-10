
import sys
import os.path
import time

import shutil

import argparse
import requests
from requests.auth import HTTPBasicAuth
import openpyxl
import re
import json
from json import dumps, loads
from jinja2 import Environment, FileSystemLoader, select_autoescape
import yaml

# Has to be located (with supporting files) in the [base]/input/scripts folder
#Call by running: python parameterize_profiles.py SDOHCC_IG_Guide_v2.0.xlsx {VSAC API Key}
# VSAC API Key used to retrieve ValueSet titles and place in to base profile MD file tables.

# Spreadsheet Constants
PARAMETERS_SPREADSHEET_NAME = 'VSAC Value Sets'
DOMAIN_CODE = 'Category code'
DOMAIN_DESCRIPTION = 'SDOH Category'

JINJA_TEMPLATE_PROFILE_FILE = 'SDOHCC_Profile_template.j2'
JINJA_PARAMETER_TABLE_TEMPLATE_FILE = 'Parameter_Table.j2'

DETAILED_PROFILE_FOLDER = '../fsh/profiles/detailed'
DETAILED_PROFILE_FOLDER = '../fsh/profiles/detailed'

PAGE_CONTENT_FOLDER = '../pagecontent'
PROFILE_MD_FILE_PREFIX = 'StructureDefinition-SDOHCC-'
PROFILE_MD_FILE_INTRO_SUFFIX = '-intro.md'
PROFILE_MD_FILE_NOTES_SUFFIX = '-notes.md'

DETAILED_PROFILE_GROUP_CODE = 'DetailedProfiles'
DETAILED_PROFILE_GROUP_NAME = 'Detailed Profiles'
DETAILED_PROFILE_GROUP_DESCRIPTION = 'The detailed profiles provides an expansion of all of the general base profiles with the specific domains bindings.'


VSAC_HOTSNAME = 'https://cts.nlm.nih.gov/fhir'


PARAMETER_TABLE_START = '\\[//\\]: # \\(Parameter Table DO NOT REMOVE\\)'
PARAMETER_TABLE_END = '\\[//\\]: # \\(End Parameter Table DO NOT REMOVE\\)'

# installs
# pip install 
#   pyyaml
#   openpyxl
#   argparse
#   json
#   jinja2
#   requests


# TODO:
# Address different resource types requirements. Need to support a default jinja template and then support specialized ones. Procedure has slicing on category.coding as opposed to category like most
# Update yaml detailed profiles
#   Add YAML group description for detailed profiles
# Would need to create examples for each and every
# !!!! Add constraints for each profile where it references a SDOHCC  profile. Will need to look at the original profile to find the reference-able profiles 
# ServiceRequest, Procedure, Goal, Condition, Observation Screening Response
# !!!! ServiceRequest may have references to resources that are outside of the domain.
# !!!! For Observation Screening Response - current Observation.code is pointing to the answer list. Observation.code should be Questions and Panels (for observation grouping)
# Grouping/Panel could represent several Categories, so we may want to separate out groupers. for groupers the 

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def main():

    file_count = 0
    issue_count = 0
    fatal_count = 0
    error_count = 0
    warning_count = 0
    info_count = 0
    
    parser = argparse.ArgumentParser(description="""SDOH Parameterized Profile Generator""")
    parser.add_argument('file_path', type=file_arg, help="Path to parameters xlsx file")    
    parser.add_argument('vsac_apikey', nargs='?', help="API Key for Value Set Authority Center (VSAC) [Needed to retrieve ValueSet titles for base profile parameter tables]")    
    args = parser.parse_args()

    file_path = args.file_path
    vsac_apikey = args.vsac_apikey
    
    
    start = time.time()    

    print(f'....processing parameters in {file_path} .....')


    try:
        shutil.rmtree(DETAILED_PROFILE_FOLDER)
        print(f"Detailed profile folder removed: {DETAILED_PROFILE_FOLDER}")
    except OSError as o:
        print(f"Detailed profile folder does not exist: {DETAILED_PROFILE_FOLDER}")

    try:
        os.makedirs(DETAILED_PROFILE_FOLDER, exist_ok=True)
        print(f"Detailed profile folder created: {DETAILED_PROFILE_FOLDER}")
    except OSError as o:
        print(f"Detailed profile folder could not be created: {DETAILED_PROFILE_FOLDER}")
        exit()

    parameters_wb = openpyxl.load_workbook(file_path)
    parameters_ws = parameters_wb[PARAMETERS_SPREADSHEET_NAME]
    #parameters_ws = parameters_wb[PARAMETERS_SPREADSHEET_NAME]
    #for row in parameters_ws.iter_rows('C{}:C{}'.format(ws.min_row,ws.max_row)):
    #    for cell in row:
    #        print cell.value
    row_index = 0

    profile_column_start = 0
    column = {}
    parameters = {}
    for domain in parameters_ws:
        if row_index == 0: # Header Row
            column_index = 0
            while column_index < parameters_ws.max_column:
                if domain[column_index].value != None:
                    profile_type = domain[column_index].value
                    column[profile_type] = column_index
                    if(profile_column_start > 0):
                        resource_info = {}
                        
                        resource_info['resource_type'] = profile_type
                        resource_info['profile_name'] = profile_type
                        resource_info['domains'] = []
                        if '.' in profile_type:
                            resource_info['resource_type'] = profile_type.split('.')[0]
                            resource_info['profile_name'] = profile_type.replace('.', '').replace('_', '')
                        
                        parameters[profile_type] = resource_info
                        
                    if domain[column_index].value == DOMAIN_CODE:
                        profile_column_start = column_index + 1

                else:
                    break
                column_index = column_index + 1
        else: # Non Header Row
            column_index = profile_column_start
            while column_index < parameters_ws.max_column:
                if domain[column_index].value != None:
                    profile_type = parameters_ws[1][column_index].value
                    if hasattr(domain[column_index], 'hyperlink') and hasattr(domain[column_index].hyperlink, 'target'):
                        category_vs = domain[column_index].hyperlink.target.replace('https://vsac.nlm.nih.gov/valueset/', 'http://cts.nlm.nih.gov/fhir/ValueSet/').replace('/expansion/Latest', '')
                        category_vs_browse = domain[column_index].hyperlink.target
                        category_vs_oid = domain[column_index].hyperlink.target.replace('https://vsac.nlm.nih.gov/valueset/', '').replace('/expansion/Latest', '')
                        category_vs_title= "None"
                        
                        if(vsac_apikey != None):
                            category_vs_title = get_vs_title(vsac_apikey, category_vs_oid)

                    else:
                        category_vs = 'None'
                    binding_details = {}
                    element = domain[column_index].value
                    if '.' in element:
                        element = element.split('.')[1]
                    
                    binding_details['domain_code'] = domain[column[DOMAIN_CODE]].value
                    binding_details['domain_title'] = domain[column[DOMAIN_DESCRIPTION]].value
                    # bindings was put in to an array in the event more than one binding change would be necessary
                    binding = {}
                    binding['element'] = element
                    binding['valueset'] = category_vs
                    binding['valueset_browse'] = category_vs_browse
                    binding['valueset_oid'] = category_vs_oid
                    binding['valueset_title'] = category_vs_title
                    
                    bindings = []
                    bindings.append(binding)
                    binding_details['bindings'] = bindings

                    parameters[profile_type]['domains'].append(binding_details)
                    
                column_index = column_index + 1
        row_index = row_index + 1
    #print(parameters)
    
    

    env = Environment(
        loader=FileSystemLoader(searchpath=os.path.abspath(os.path.dirname(os.path.realpath(__file__)))),
        autoescape=select_autoescape(['html', 'xml', 'xhtml', 'j2', 'md'])
    )

    
    with open('../../sushi-config.yaml', 'r') as file:
        sushi_config = yaml.safe_load(file)
        groups = sushi_config['groups']
        #if DETAILED_PROFILE_GROUP_CODE in groups:
        #    del groups[DETAILED_PROFILE_GROUP_CODE]
         
        groups[DETAILED_PROFILE_GROUP_CODE] = {}
        groups[DETAILED_PROFILE_GROUP_CODE]['name'] = DETAILED_PROFILE_GROUP_NAME
        groups[DETAILED_PROFILE_GROUP_CODE]['description'] = DETAILED_PROFILE_GROUP_DESCRIPTION
        
        #print(sushi_config['groups'])

    
        # Loop through each base Profile Type
        detailed_profile_list = []
        for profile_type in parameters:
            profile = parameters[profile_type]


            # Render table data in intro or notes md file
            profile_specific_template = JINJA_PARAMETER_TABLE_TEMPLATE_FILE.replace(".j2", "_" + profile['profile_name'] + ".j2")
            resource_specific_template = JINJA_PARAMETER_TABLE_TEMPLATE_FILE.replace(".j2", "_" + profile['resource_type'] + ".j2")
            # Some resources have unique requirements, e.g. Procedure slicing on category is category.code
            if os.path.isfile(profile_specific_template):
                template = env.get_template(profile_specific_template)
            if os.path.isfile(resource_specific_template):
                template = env.get_template(resource_specific_template)
            else:
                template = env.get_template(JINJA_PARAMETER_TABLE_TEMPLATE_FILE)
            
            rendered = template.render(resource_type=profile['resource_type'],profile_name=profile['profile_name'],profile_title=profile_type.replace('.', ' ').replace('_', ' '), domains=profile['domains'])
            #print(rendered)

            update_parameter_table(profile['profile_name'], rendered)
            


            
            profile_specific_template = JINJA_TEMPLATE_PROFILE_FILE.replace(".j2", "_" + profile['profile_name'] + ".j2")
            resource_specific_template = JINJA_TEMPLATE_PROFILE_FILE.replace(".j2", "_" + profile['resource_type'] + ".j2")
            # Some resources have unique requirements, e.g. Procedure slicing on category is category.code
            if os.path.isfile(profile_specific_template):
                template = env.get_template(profile_specific_template)
            if os.path.isfile(resource_specific_template):
                template = env.get_template(resource_specific_template)
            else:
                template = env.get_template(JINJA_TEMPLATE_PROFILE_FILE)
            
            for domain in profile['domains']:
                rendered = template.render(resource_type=profile['resource_type'],profile_name=profile['profile_name'],profile_title=profile_type.replace('.', ' ').replace('_', ' '), domain=domain)
                detailed_profile_file_name = DETAILED_PROFILE_FOLDER + "/" + "SDOHCC" + profile['profile_name'] + "_" + domain['domain_code'] + '.fsh'
                #print("Writing profile SDOHCC" + profile['profile_name'] + "-" + domain['domain_code'] + " to file: " + detailed_profile_file_name)
                with open(detailed_profile_file_name, 'w') as file:
                    file.write(rendered)
                detailed_profile_list.append("SDOHCC" + profile['profile_name'] + "_" + str(domain['domain_code']).replace("-", "_"))
        
        groups[DETAILED_PROFILE_GROUP_CODE]['resources'] = detailed_profile_list
                #detailed_profile_file_name.write_text(rendered)
                
                #print("")
                #print(rendered)
    #print(sushi_config)
    with open('../../sushi-config.yaml', 'w') as file:
        yaml.dump(sushi_config, file, sort_keys=False, default_flow_style=False)
        #yaml.dump(sushi_config, sort_keys=False, default_flow_style=False)
#template.render(cs=cs, path_map=path_map, pname_map=pname_map, purl_map=purl_map, sp_map=sp_map, 
#                       csname_map=csname_map, csurl_map=csurl_map, igname_map=igname_map, igurl_map=igurl_map)

    #tempPath = Path.cwd() / "test.html"
    #tempPath.write_text(rendered)
    
    end = time.time()
    print("Execution time (in seconds)", end - start)


def get_vs_title(vsac_apikey, category_vs_oid):
    print("Retrieving ValueSet " + category_vs_oid + " from VSAC.")
    get_try = 1
    while get_try <= 5:
        try:
            response = requests.get(VSAC_HOTSNAME + '/res/ValueSet/' + category_vs_oid, auth=('apikey', vsac_apikey), timeout=10)
            if(response.status_code == 200):
                json_data = json.loads(response.content)
                print("Successfully ValueSet " + category_vs_oid + " from VSAC.")
                return json_data['title']
            else:
                print(bcolors.BOLD + bcolors.FAIL + "Unable to retrieve ValueSet " + category_vs_oid + " from VSAC. The MD file will have to be edited manually." + bcolors.ENDC)
                print(bcolors.BOLD + bcolors.FAIL + "Code: " + str(response.status_code) + " Reason: " + response.reason + bcolors.ENDC)
        except requests.exceptions.Timeout:
            print("Timed out")
        get_try += 1

    return "Unable to retrieve ValueSet Title from VSAC"

def update_parameter_table(base_profile_id, table):
    md_file = PAGE_CONTENT_FOLDER + '/' + PROFILE_MD_FILE_PREFIX + base_profile_id + PROFILE_MD_FILE_INTRO_SUFFIX
    if os.path.isfile(md_file):
        print("File Exists! " + md_file)
        with open(md_file, 'r+') as file:
            # read all content from a file using read()
            content = file.read()
            # check if string present or not
            match = re.search(PARAMETER_TABLE_START , content)  
            if match != None:
                #print(content)
                print('!!!! REPLACE TABLE EXISTS!!!!!')
                
                content = re.sub(PARAMETER_TABLE_START + '.*?' + PARAMETER_TABLE_END, PARAMETER_TABLE_START.replace('\\', '') + '\n\n' + table + '\n\n' + PARAMETER_TABLE_END.replace('\\', ''), content, flags=re.DOTALL)
                #print(content)
                #file.write(content)
            
            else:
                content = ''
                print('------REPLACE TABLE DOES NOT EXIST------')
        if(content != ''):
            with open(md_file, 'w') as file:
                print(content)
                file.write(content)
                
    md_file = PAGE_CONTENT_FOLDER + '/' + PROFILE_MD_FILE_PREFIX + base_profile_id + PROFILE_MD_FILE_NOTES_SUFFIX
    if os.path.isfile(md_file):
        print("File Exists! " + md_file)


def getValue(data, title):
    #if((title in data) and (data[title])):
    if((str(title) in data) and (str(data[title]).upper() != 'NAN')):
        return str(data[title])
    else:
        return ""

def file_arg(string):
    if os.path.isfile(string):
        return string
    else:
        raise argparse.ArgumentError(f'D{string}, is not a found file')

main()
